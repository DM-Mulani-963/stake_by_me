"""Excel file generator from processed data"""

import logging
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Optional
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

logger = logging.getLogger(__name__)


class ExcelGenerator:
    """Generate Excel files with structured registration data"""
    
    def __init__(self, output_folder: str = "./output"):
        self.output_folder = Path(output_folder)
        self.output_folder.mkdir(parents=True, exist_ok=True)
        
        # Define column structure
        self.columns = [
            "email",
            "username",
            "password",
            "dateofbirth",
            "phonenumber",
            "firstname",
            "lastname",
            "country",
            "place_of_birth",
            "residential_address",
            "city",
            "postal_code",
            "occupation_industry",
            "occupation_field",
            "occupation_experience",
            # Additional info
            "full_name",
            "blood_group",
            "license_number",
            "state",
            "source_file",
        ]
        
        # Column headers (display names)
        self.column_headers = {
            "email": "Email",
            "username": "Username",
            "password": "Password",
            "dateofbirth": "Date of Birth",
            "phonenumber": "Phone Number",
            "firstname": "First Name",
            "lastname": "Last Name",
            "country": "Country",
            "place_of_birth": "Place of Birth",
            "residential_address": "Residential Address",
            "city": "City",
            "postal_code": "Postal Code",
            "occupation_industry": "Occupation Industry",
            "occupation_field": "Occupation Field",
            "occupation_experience": "Occupation Experience",
            "full_name": "Full Name (License)",
            "blood_group": "Blood Group",
            "license_number": "License Number",
            "state": "State",
            "source_file": "Source JSON File",
        }
    
    def create_excel(self, records: List[Dict], filename: Optional[str] = None) -> Path:
        """
        Create Excel file from records
        
        Args:
            records: List of record dictionaries
            filename: Optional custom filename, otherwise auto-generated
            
        Returns:
            Path to created Excel file
        """
        if not records:
            logger.warning("No records to write to Excel")
            return None
        
        # Generate filename if not provided
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"registration_data_{timestamp}.xlsx"
        
        # Ensure .xlsx extension
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        
        filepath = self.output_folder / filename
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Registration Data"
        
        # Style definitions
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        cell_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        
        # Write headers
        for col_idx, col_key in enumerate(self.columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = self.column_headers.get(col_key, col_key)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = cell_border
        
        # Write data rows
        for row_idx, record in enumerate(records, start=2):
            for col_idx, col_key in enumerate(self.columns, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                
                # Get value with fallback
                value = record.get(col_key, record.get(f"_{col_key}", ""))
                
                # Special handling for source file
                if col_key == "source_file":
                    value = record.get("_source_file", "")
                
                cell.value = str(value) if value is not None else ""
                cell.border = cell_border
                
                # Center align for specific columns
                if col_key in ["dateofbirth", "country", "state", "blood_group"]:
                    cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths
        for col_idx, col_key in enumerate(self.columns, start=1):
            column_letter = openpyxl.utils.get_column_letter(col_idx)
            
            # Set minimum width based on header
            header_length = len(self.column_headers.get(col_key, col_key))
            max_length = max(header_length, 15)
            
            # Check data length (sample first few rows)
            for row_idx in range(2, min(len(records) + 2, 12)):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            
            # Cap maximum width
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
        # Add data validation for specific columns (optional enhancement)
        self._add_data_validation(ws, len(records))
        
        # Save workbook
        try:
            wb.save(filepath)
            logger.info(f"Excel file created successfully: {filepath}")
            logger.info(f"Total rows: {len(records)}")
            return filepath
        except Exception as e:
            logger.error(f"Error saving Excel file: {e}")
            raise
    
    def _add_data_validation(self, ws, num_records: int):
        """Add data validation rules to specific columns"""
        # This is optional - adds dropdown validation for certain fields
        # Currently just demonstrates the capability
        
        # Example: Add validation for occupation experience
        exp_validation = DataValidation(
            type="list",
            formula1='"0-2 years,3-5 years,6-10 years,10+ years"',
            allow_blank=True
        )
        ws.add_data_validation(exp_validation)
        
        # Apply to occupation experience column (if it exists)
        try:
            exp_col_idx = self.columns.index("occupation_experience") + 1
            exp_col_letter = openpyxl.utils.get_column_letter(exp_col_idx)
            exp_validation.add(f"{exp_col_letter}2:{exp_col_letter}{num_records + 1}")
        except ValueError:
            pass  # Column not found
    
    def create_summary_sheet(self, wb, records: List[Dict]):
        """Create a summary sheet with statistics (optional enhancement)"""
        ws = wb.create_sheet("Summary")
        
        # Add summary statistics
        ws['A1'] = "Summary Statistics"
        ws['A1'].font = Font(bold=True, size=14)
        
        ws['A3'] = "Total Records:"
        ws['B3'] = len(records)
        
        ws['A4'] = "Generated On:"
        ws['B4'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Count by state
        ws['A6'] = "Records by State:"
        row = 7
        state_counts = {}
        for record in records:
            state = record.get("state", "Unknown")
            state_counts[state] = state_counts.get(state, 0) + 1
        
        for state, count in sorted(state_counts.items()):
            ws[f'A{row}'] = state
            ws[f'B{row}'] = count
            row += 1


# Convenience function
def generate_excel(records: List[Dict], filename: Optional[str] = None, output_folder: str = "./output") -> Path:
    """Generate Excel file from records"""
    generator = ExcelGenerator(output_folder)
    return generator.create_excel(records, filename)
